"""
설명: AI Agent가 사용하는 통합 리포트 생성 도구 (All-in-One, Updated Logic)

[Tools List]
1. generate_monthly_purchase_closing_report: 월말 마감 (Summary/TopItems/Detail/Error)
2. generate_po_status_report: 발주 현황 (최근 2년, 제품유형별 시트 분할)
3. generate_supplier_evaluation_report: 공급업체 평가 (부적합/납기준수율) - Logic Updated
4. generate_excel_report: 범용 엑셀 생성
"""

import pandas as pd
import numpy as np
import os
import re
import json
from datetime import datetime
from langchain_core.tools import tool
# 서식 적용을 위한 openpyxl 모듈
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

from .shared import DB, OUTPUT_DIR


# =============================================================================
# Helper Utilities
# =============================================================================
def _normalize_id(x):
    """ID 정규화 (공백/소수점/앞자리 0 제거)"""
    if pd.isna(x): return None
    s = str(x).strip()
    if s.lower() == 'nan': return None
    if s.endswith('.0'): s = s[:-2]
    if not s: return None
    return s.lstrip('0')


def _safe_save_excel(df: pd.DataFrame, file_path: str, sheet_name: str = "Report") -> str:
    """안전한 엑셀 저장 (실패 시 CSV Fallback)"""
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        return f"파일 생성 완료: {file_path}"
    except Exception as e:
        csv_path = file_path.replace('.xlsx', '.csv')
        df.to_csv(csv_path, index=False)
        return f"엑셀 생성 실패({e}) -> CSV 저장됨: {csv_path}"


def _apply_excel_formatting(ws):
    """
    [서식 적용 헬퍼]
    1) 헤더: 배경 #D9D9D9, 텍스트 굵게, 중앙 정렬
    2) 내용: 숫자 형식(#,##0), 모든 셀 테두리 적용
    """
    # 스타일 정의
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    max_row = ws.max_row
    max_col = ws.max_column

    # 1. 헤더 서식 적용 (첫 번째 행)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 2. 본문 서식 적용 (2행부터)
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            # 숫자 형식 적용
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    # 3. 열 너비 자동 조정 (간략 버전)
    for column_cells in ws.columns:
        length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 4, 50)


# =============================================================================
# 1. 범용 엑셀 생성 도구
# =============================================================================
@tool
def generate_excel_report(data_json: str, filename: str = "report.xlsx") -> str:
    """
    [범용 파일 생성] AI가 분석한 데이터(JSON)를 엑셀 파일로 저장합니다.
    """
    try:
        data = json.loads(data_json) if isinstance(data_json, str) else data_json
        df = pd.DataFrame(data)
        if not filename.endswith('.xlsx'): filename += '.xlsx'
        file_path = os.path.join(OUTPUT_DIR, filename)

        # 범용 툴도 서식 적용을 위해 별도 처리
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="Report")
                _apply_excel_formatting(writer.sheets["Report"])
            return f"파일 생성 완료: {file_path}"
        except Exception as e:
            return _safe_save_excel(df, file_path)  # Fallback

    except Exception as e:
        return f"파일 생성 실패: {str(e)}"


# =============================================================================
# 2. 월말 구매 마감 리포트
# =============================================================================
@tool
def generate_monthly_purchase_closing_report() -> str:
    """
    [월말 구매 마감 리포트 생성]
    구매 상세 내역을 분석하여 4개 시트(Summary, Top_Items, Detail, Error_Log)가 포함된 엑셀을 생성합니다.
    **서식 적용: 헤더(회색/굵게), 테두리, 숫자 형식**
    """
    print(">>> [Tool] 월말 구매 마감 리포트 생성 시작")
    try:
        # 데이터 로드
        txn_df = DB.get('purchase_transaction_history', pd.DataFrame()).copy()
        prod_df = DB.get('product', pd.DataFrame()).copy()

        if txn_df.empty: return "데이터 오류: 구매 상세 내역이 비어있습니다."

        # 전처리
        if 'receipt_date' in txn_df.columns:
            txn_df['receipt_date'] = pd.to_datetime(txn_df['receipt_date'], errors='coerce')

        if 'movement_type' in txn_df.columns:
            txn_df['movement_type'] = txn_df['movement_type'].astype(str).str.replace(r'\.0$', '',
                                                                                      regex=True).str.strip()
            # 101(입고), 102(취소)만 대상
            txn_df = txn_df[txn_df['movement_type'].isin(['101', '102'])].copy()

        # ID 정규화 및 병합
        txn_df['product_id'] = txn_df['product_id'].apply(_normalize_id)
        prod_df['product_id'] = prod_df['product_id'].apply(_normalize_id)

        prod_unique = prod_df.drop_duplicates(subset=['product_id'], keep='first')
        merged_df = txn_df.merge(prod_unique[['product_id', 'product_type', 'description']], on='product_id',
                                 how='left')

        # 금액 계산 (원화)
        target_col = 'total_received_value_krw'
        if target_col not in merged_df.columns:
            merged_df[target_col] = 0
        else:
            merged_df[target_col] = pd.to_numeric(merged_df[target_col], errors='coerce').fillna(0)
            if 'movement_type' in merged_df.columns:
                mask_cancel = merged_df['movement_type'] == '102'
                merged_df.loc[mask_cancel, target_col] = -np.abs(merged_df.loc[mask_cancel, target_col])

        # 자재 유형 정제
        merged_df['product_type'] = merged_df['product_type'].astype(str).str.strip().str.upper()
        merged_df.loc[merged_df['product_type'].isin(['NAN', 'NONE', '']), 'product_type'] = 'UNCLASSIFIED'

        # 기준 연월
        valid_dates = merged_df['receipt_date'].dropna()
        if not valid_dates.empty:
            max_date = valid_dates.max()
            curr_year, curr_month = max_date.year, max_date.month
        else:
            now = datetime.now()
            curr_year, curr_month = now.year, now.month

        # Sheet 1: Summary 로직
        def get_quarter_str(year, month):
            return f"{year} {(month - 1) // 3 + 1}Q"

        if curr_month == 1:
            prev_month_date = datetime(curr_year - 1, 12, 1)
            ref1_y, ref1_m = curr_year - 1, 6;
            ref2_y, ref2_m = curr_year - 1, 9
        else:
            prev_month_date = datetime(curr_year, curr_month - 1, 1)
            if 1 < curr_month <= 3:
                ref1_y, ref1_m = curr_year - 1, 9; ref2_y, ref2_m = curr_year - 1, 12
            elif 3 < curr_month <= 6:
                ref1_y, ref1_m = curr_year - 1, 12; ref2_y, ref2_m = curr_year, 3
            elif 6 < curr_month <= 9:
                ref1_y, ref1_m = curr_year, 3; ref2_y, ref2_m = curr_year, 6
            else:
                ref1_y, ref1_m = curr_year, 6; ref2_y, ref2_m = curr_year, 9

        prev_month_y, prev_month_m = prev_month_date.year, prev_month_date.month
        col_ref1 = get_quarter_str(ref1_y, ref1_m)
        col_ref2 = get_quarter_str(ref2_y, ref2_m)

        def calculate_sum(year, month, p_type, curr_cond):
            cond = (merged_df['product_type'] == p_type)
            if month != 0:
                cond &= (merged_df['receipt_date'].dt.year == year) & (merged_df['receipt_date'].dt.month == month)
            else:
                cond &= (merged_df['receipt_date'].dt.year == year)

            if 'order_currency' in merged_df.columns:
                curr_series = merged_df['order_currency'].fillna('').astype(str).str.strip().str.upper()
                if curr_cond == 'KRW':
                    cond &= (curr_series == 'KRW')
                elif curr_cond == 'NON-KRW':
                    cond &= (curr_series != 'KRW')
            return float(merged_df.loc[cond, target_col].sum())

        roh1_krw_curr = calculate_sum(curr_year, curr_month, 'ROH1', 'KRW')
        roh1_krw_prev = calculate_sum(prev_month_y, prev_month_m, 'ROH1', 'KRW')
        roh1_krw_last = calculate_sum(curr_year - 1, 0, 'ROH1', 'KRW')
        roh1_krw_ref1 = calculate_sum(ref1_y, ref1_m, 'ROH1', 'KRW')
        roh1_krw_ref2 = calculate_sum(ref2_y, ref2_m, 'ROH1', 'KRW')

        roh1_for_curr = calculate_sum(curr_year, curr_month, 'ROH1', 'NON-KRW')
        roh1_for_prev = calculate_sum(prev_month_y, prev_month_m, 'ROH1', 'NON-KRW')
        roh1_for_last = calculate_sum(curr_year - 1, 0, 'ROH1', 'NON-KRW')
        roh1_for_ref1 = calculate_sum(ref1_y, ref1_m, 'ROH1', 'NON-KRW')
        roh1_for_ref2 = calculate_sum(ref2_y, ref2_m, 'ROH1', 'NON-KRW')

        roh2_krw_curr = calculate_sum(curr_year, curr_month, 'ROH2', 'KRW')
        roh2_krw_prev = calculate_sum(prev_month_y, prev_month_m, 'ROH2', 'KRW')
        roh2_krw_last = calculate_sum(curr_year - 1, 0, 'ROH2', 'KRW')
        roh2_krw_ref1 = calculate_sum(ref1_y, ref1_m, 'ROH2', 'KRW')
        roh2_krw_ref2 = calculate_sum(ref2_y, ref2_m, 'ROH2', 'KRW')

        summary_rows = [
            {"구분": "내자 원료", "작년 총합": roh1_krw_last, col_ref1: roh1_krw_ref1, col_ref2: roh1_krw_ref2,
             "전월 실적": roh1_krw_prev, "당월 실적": roh1_krw_curr, "전월 대비 증감": roh1_krw_curr - roh1_krw_prev},
            {"구분": "외자 원료 (KRW)", "작년 총합": roh1_for_last, col_ref1: roh1_for_ref1, col_ref2: roh1_for_ref2,
             "전월 실적": roh1_for_prev, "당월 실적": roh1_for_curr, "전월 대비 증감": roh1_for_curr - roh1_for_prev},
            {"구분": "내자 자재", "작년 총합": roh2_krw_last, col_ref1: roh2_krw_ref1, col_ref2: roh2_krw_ref2,
             "전월 실적": roh2_krw_prev, "당월 실적": roh2_krw_curr, "전월 대비 증감": roh2_krw_curr - roh2_krw_prev},
        ]

        def sum_rows(title, *rows):
            result = {"구분": title}
            keys = [k for k in rows[0].keys() if k != "구분"]
            for k in keys: result[k] = sum(row[k] for row in rows)
            return result

        summary_rows.append(sum_rows("원료 합계 (내자+외자)", summary_rows[0], summary_rows[1]))
        summary_rows.append(sum_rows("내자 합계 (원료+자재)", summary_rows[0], summary_rows[2]))
        summary_rows.append(sum_rows("전체 합계 (Total)", summary_rows[0], summary_rows[1], summary_rows[2]))
        df_summary = pd.DataFrame(summary_rows)

        # Sheet 2: Top_Items
        current_month_df = merged_df[
            (merged_df['receipt_date'].dt.year == curr_year) &
            (merged_df['receipt_date'].dt.month == curr_month)
            ].copy()

        top_items_list = []
        df_top_items = pd.DataFrame()

        if not current_month_df.empty:
            if 'po_id' not in current_month_df.columns: current_month_df['po_id'] = '-'
            grouped = current_month_df.groupby(['product_id', 'product_type']).agg({
                target_col: 'sum', 'description': 'first',
                'po_id': lambda x: ', '.join(sorted(x.dropna().astype(str).unique()))[:100]
            }).reset_index()

            roh1_top = grouped[grouped['product_type'] == 'ROH1'].nlargest(1, target_col)
            if not roh1_top.empty:
                row = roh1_top.iloc[0].to_dict()
                row['구분'] = '원료(ROH1) 최다 지출'
                top_items_list.append(row)

            roh2_top = grouped[grouped['product_type'] == 'ROH2'].nlargest(3, target_col)
            if not roh2_top.empty:
                rank = 1
                for _, row in roh2_top.iterrows():
                    r_dict = row.to_dict()
                    r_dict['구분'] = f'자재(ROH2) Top {rank}'
                    top_items_list.append(r_dict)
                    rank += 1

            if top_items_list:
                df_top_items = pd.DataFrame(top_items_list)
                target_cols = ['구분', 'product_id', 'description', target_col, 'po_id']
                existing_cols = [c for c in target_cols if c in df_top_items.columns]
                df_top_items = df_top_items[existing_cols]
                rename_map = {'product_id': '자재 코드', 'description': '자재명', target_col: '금액', 'po_id': '관련 PO'}
                df_top_items = df_top_items.rename(columns=rename_map)

        # Excel Save
        df_detail = current_month_df.copy()
        error_mask = (merged_df['product_type'] == 'UNCLASSIFIED') | (merged_df['description'].isna())
        df_error = merged_df[error_mask].copy()

        filename = f"Monthly_Closing_Report_{curr_year}_{curr_month:02d}.xlsx"
        file_path = os.path.join(OUTPUT_DIR, filename)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 1. Summary
            df_summary.to_excel(writer, sheet_name="Summary", index=False)
            _apply_excel_formatting(writer.sheets["Summary"])

            # 2. Top_Items
            if not df_top_items.empty:
                df_top_items.to_excel(writer, sheet_name="Top_Items", index=False)
                _apply_excel_formatting(writer.sheets["Top_Items"])

            # 3. Detail
            df_detail.to_excel(writer, sheet_name="Detail_Data", index=False)
            _apply_excel_formatting(writer.sheets["Detail_Data"])

            # 4. Error
            if not df_error.empty:
                df_error.to_excel(writer, sheet_name="Error_Log", index=False)
                _apply_excel_formatting(writer.sheets["Error_Log"])
            else:
                pd.DataFrame({'Status': ['No Errors Found']}).to_excel(writer, sheet_name="Error_Log", index=False)
                _apply_excel_formatting(writer.sheets["Error_Log"])

        # Markdown Preview
        summary_display = df_summary.copy()
        for col in summary_display.select_dtypes(include=['number']).columns:
            summary_display[col] = summary_display[col].apply(lambda x: f"{x:,.0f}")

        return f"### 📊 {curr_year}년 {curr_month}월 마감 요약 (미리보기)\n{summary_display.to_markdown(index=False)}\n\n✅ **리포트 생성 완료**: `{file_path}`"

    except Exception as e:
        return f"오류 발생: {str(e)}"


# =============================================================================
# 3. 발주 현황 공유 파일 생성
# =============================================================================
@tool
def generate_po_status_report() -> str:
    """
    [발주 현황 공유 파일 생성]
    최근 2년 데이터를 조회하고, '제품 유형'별로 시트를 분할하여 저장합니다.
    **서식 적용: 헤더(회색/굵게), 테두리, 숫자 형식**
    """
    print(">>> [Tool] 발주 현황 리포트 생성 시작")
    try:
        po_df = DB.get('purchase_order', pd.DataFrame()).copy()
        vendor_df = DB.get('vendor_info_record', pd.DataFrame()).copy()
        prod_df = DB.get('product', pd.DataFrame()).copy()

        if po_df.empty: return "데이터 오류: 구매오더 데이터가 없습니다."

        po_df['vendor_id'] = po_df['vendor_id'].apply(_normalize_id)
        po_df['product_id'] = po_df['product_id'].apply(_normalize_id)
        vendor_df['vendor_id'] = vendor_df['vendor_id'].apply(_normalize_id)
        prod_df['product_id'] = prod_df['product_id'].apply(_normalize_id)

        if 'po_date' in po_df.columns: po_df['po_date'] = pd.to_datetime(po_df['po_date'], errors='coerce')

        today = datetime.now()
        two_years_ago = today - pd.DateOffset(years=2)
        filtered_po = po_df[(po_df['po_date'] >= two_years_ago) & (po_df['po_date'] <= today)].copy()

        merged_df = filtered_po.merge(vendor_df[['vendor_id', 'vendor_name']].drop_duplicates(), on='vendor_id',
                                      how='left')
        merged_df = merged_df.merge(prod_df[['product_id', 'description', 'product_type']].drop_duplicates(),
                                    on='product_id', how='left')
        merged_df.loc[merged_df['product_type'].isna(), 'product_type'] = 'Unclassified'

        file_date = today.strftime("%Y%m%d")
        filename = f"PO_Status_Share_{file_date}.xlsx"
        file_path = os.path.join(OUTPUT_DIR, filename)

        # Markdown 요약
        summary = merged_df['product_type'].value_counts().reset_index()
        summary.columns = ['제품 유형', '발주 건수']
        summary_md = summary.to_markdown(index=False)

        # Excel Save with Formatting
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for p_type in merged_df['product_type'].unique():
                sheet_name = re.sub(r'[\\/*?:\[\]]', '', str(p_type))[:30] or "Etc"
                # 시트 저장
                merged_df[merged_df['product_type'] == p_type].to_excel(writer, sheet_name=sheet_name, index=False)
                # 서식 적용
                _apply_excel_formatting(writer.sheets[sheet_name])

        return f"### 📂 발주 현황 요약 (최근 2년)\n{summary_md}\n\n✅ **파일 생성 완료**: `{file_path}`"
    except Exception as e:
        return f"실패: {str(e)}"


# =============================================================================
# 4. 공급업체 평가 양식 생성
# =============================================================================
@tool
def generate_supplier_evaluation_report() -> str:
    """
    [공급업체 평가 양식 생성]
    1) Batch No -> Good Receipt -> Vendor ID -> Vendor Name 매핑으로 정확도 향상.
    2) 납품일(LT) = Good Receipt(Delivery/Receipt Date) - PO Date 로 계산하여 실소요기간 반영.
    **서식 적용: 헤더(회색/굵게), 테두리, 숫자 형식**
    """
    print(">>> [Tool] 공급업체 평가 양식 생성 시작")
    try:
        gr_df = DB.get('good_receipt', pd.DataFrame()).copy()
        po_df = DB.get('purchase_order', pd.DataFrame()).copy()
        nc_df = DB.get('non_conformance', pd.DataFrame()).copy()
        vendor_df = DB.get('vendor_info_record', pd.DataFrame()).copy()
        prod_df = DB.get('product', pd.DataFrame()).copy()

        if gr_df.empty: return "오류: 입고 내역이 없습니다."

        # 정규화
        for df in [gr_df, po_df, nc_df, vendor_df, prod_df]:
            for col in ['batch_no', 'po_id', 'product_id', 'vendor_id']:
                if col in df.columns: df[col] = df[col].apply(_normalize_id)

        # 날짜 변환 (PO delivery_date 포함)
        for df, col in [(gr_df, 'receipt_date'), (po_df, 'po_date'), (po_df, 'delivery_date')]:
            if col in df.columns: df[col] = pd.to_datetime(df[col], errors='coerce')

        nc_batch_set = set(nc_df['batch_no'].dropna().unique()) if not nc_df.empty else set()

        # GR -> PO -> Vendor 순차 병합
        gr_base = gr_df.dropna(subset=['batch_no']).drop_duplicates(subset=['batch_no']).copy()
        gr_base['is_non_conformance'] = gr_base['batch_no'].apply(lambda x: "부적합" if x in nc_batch_set else "적합")

        # 1. GR + PO (delivery_date 가져오기)
        # PO에서 delivery_date 컬럼을 명시적으로 포함하여 병합
        merged = gr_base.merge(po_df[['po_id', 'vendor_id', 'po_date', 'delivery_date']].drop_duplicates('po_id'),
                               on='po_id', how='left')

        # 2. + Vendor Info
        merged = merged.merge(vendor_df[['vendor_id', 'vendor_name']].drop_duplicates('vendor_id'), on='vendor_id',
                              how='left')

        # 3. + Product Info
        merged = merged.merge(prod_df[['product_id', 'description']].drop_duplicates('product_id'), on='product_id',
                              how='left')

        # 납기일(LT) 계산: delivery_date(1순위) -> receipt_date(2순위)
        # PO 테이블에 있는 delivery_date를 우선적으로 사용
        target_date_col = None
        if 'delivery_date' in merged.columns:
            target_date_col = 'delivery_date'
        elif 'receipt_date' in merged.columns:
            target_date_col = 'receipt_date'

        # target 컬럼이 존재하고 po_date가 있을 때만 계산
        if target_date_col and 'po_date' in merged.columns:
            merged['delivery_lt'] = (merged[target_date_col] - merged['po_date']).dt.days.fillna(0)
        else:
            merged['delivery_lt'] = 0

        final_cols = {
            'vendor_name': '업체명', 'product_id': '품목코드', 'description': '품목명',
            'batch_no': '성적번호', 'delivery_lt': '납품 소요일(LT)', 'is_non_conformance': '부적합 여부'
        }
        for c in ['부적합 사유', '납품 지연 사유', '지연 통보 선제성', '납품 기준 준수']: merged[c] = ""

        res = merged.rename(columns=final_cols)
        cols = list(final_cols.values()) + ['부적합 사유', '납품 지연 사유', '지연 통보 선제성', '납품 기준 준수']
        res = res[[c for c in cols if c in res.columns]].drop_duplicates()

        filename = f"Supplier_Evaluation_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path = os.path.join(OUTPUT_DIR, filename)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            res.to_excel(writer, index=False, sheet_name='평가양식')
            _apply_excel_formatting(writer.sheets['평가양식'])

        total_cnt = len(res)
        nc_cnt = len(res[res['부적합 여부'] == '부적합'])
        avg_lt = res['납품 소요일(LT)'].mean() if '납품 소요일(LT)' in res.columns else 0

        summary_md = f"""
| 항목 | 값 |
|---|---|
| 평가 대상 건수 | {total_cnt:,}건 |
| 부적합 발생 | {nc_cnt:,}건 |
| 평균 납품 소요일 | {avg_lt:.1f}일 |
"""
        return f"### 📊 공급업체 평가 요약\n{summary_md}\n\n✅ **평가 양식 생성 완료**: `{file_path}`"
    except Exception as e:
        return f"실패: {str(e)}"