import datetime

import pytest
try:
    import openpyxl
except Exception:
    openpyxl = None  # type: ignore
    pytest.skip("openpyxl is not installed; skipping merged_cell_phasing", allow_module_level=True)
import pandas as pd
import re

def parse_production_schedule(file_path):
    # 엑셀 파일 로드 (data_only=True로 수식이 아닌 값을 가져옴)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    all_plans = []

    # 시트 순회
    for sheet_name in wb.sheetnames:
        # 시트 이름이 YYYYMM 형식이 아니면 건너뛰기 (필요 시 로직 추가)
        if not re.match(r'\d{6}', sheet_name): continue
        if not re.match(r'(2025\d+)', sheet_name): continue

        ws = wb[sheet_name]

        # 메타 데이터 추출 (E3 셀: 'datetime.datetime(2025, 1, 1, 0, 0)'
        header_val = ws['E3'].value
        if not header_val:
            continue
        month, year = header_val.month, header_val.year

        # 그리드 순회 (5주 ~ 6주 가정
        # E열(5) ~ K열(11) : 일 ~ 토
        cols_range = range(5, 12)

        # 첫 번째 주 일자 행은 5행. 이후 6행씩 증가 (5, 11, 17, 23, 29, 35)
        # 최대 6주까지 체크
        for week_idx in range(6):
            date_row = 5 + (week_idx * 6)

            # 행 정의 분리
            semi_rows = [date_row + 1, date_row + 2]  # 반제품: +1, +2
            pack_rows = [date_row + 3, date_row + 4]  # 포장: +3, +4

            # 해당 주의 일자가 없으면(빈 행) 루프 종료
            if not ws.cell(row=date_row, column=5).value and not ws.cell(row=date_row, column=11).value:
                break

            # 반제품 생산 계획 파싱
            for r in semi_rows:
                # E열부터 K열까지 순회
                # *중요*: 병합된 셀은 좌측 상단 셀만 값을 가지고, 나머지는 None임.
                # 따라서 값을 가진 셀만 찾아서 처리.
                for c in cols_range:
                    cell = ws.cell(row=r, column=c)
                    cell_value = cell.value

                    if not cell_value:
                        continue

                    # 텍스트 파싱 전처리 (공백 제거 등)
                    text = str(cell_value).strip()
                    if "생산" not in text:
                        continue
                    start_date, end_date = get_merged_date(ws, cell, date_row, c, year, month)
                    if not start_date: continue

                    # 생산                     리터럴 '생산' 문자열과 매칭
                    # \s+                     공백 문자(스페이스, 탭 등) 1개 이상 (필수 구분자)
                    # (#[\d-]+)               [Group 1] 일련번호: '#'으로 시작하고 숫자나 하이픈(-)이 옴 (예: #1-1)
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # ([가-힣a-zA-Z\s\/]+?)     [Group 2] 국가/유형: 한글, 영문, 공백, 슬래시(/) 포함 (예: 유럽 / US)
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # (\d+)U?                  [Group 3] 수량: 숫자+U 형태. 포장은 수량이 생략될 수 있어 옵션 처리.
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # (\(.* ?\))?             [Group 4] (옵션) 설명: 괄호 (...)로 묶인 추가 정보. '?'는 없어도 된다는 뜻.
                    # (X[\d\s,a-fA-F]+)?      [Group 5] (옵션) 배치번호: 'X'로 시작하며 숫자, 공백, 쉼표, Hex코드 포함

                    pattern = r"생산\s+(#[\d-]+)\s*([가-힣a-zA-Z\s\/]+?)\s*(\d+)U?\s*(\(.*?\))?\s*(X[\d\s,a-fA-F]+)?"
                    p_match = re.search(pattern, text)

                    if p_match:
                        serial_no = p_match.group(1)  # #1-1
                        country = p_match.group(2)  # 유럽
                        qty = p_match.group(3)  # 50
                        description = p_match.group(4)
                        batch_no = p_match.group(5) # X12345

                        plan_data = {
                            "계획연도": year,
                            "일련번호": serial_no,
                            "유형": "반제품",  # 고정값
                            "국가": country,
                            "단위(U)": qty,
                            "설명": description,
                            "시작일": start_date,
                            "종료일": end_date,
                            "배치번호": batch_no,
                        }
                        all_plans.append(plan_data)

            # ====================================================
            # 2. 포장 계획 파싱 (신규 추가)
            # ====================================================
            for r in pack_rows:
                for c in cols_range:
                    cell = ws.cell(row=r, column=c)
                    cell_value = cell.value
                    if not cell_value: continue

                    text = str(cell_value).strip()
                    # 포장 계획 식별자 확인
                    if "포장#" not in text: continue

                    # 날짜 계산 (반제품과 동일 로직 사용)
                    start_date, end_date = get_merged_date(ws, cell, date_row, c, year, month)
                    if not start_date: continue

                    # 포장                     [Group 1] 리터럴 '포장'
                    # #                       리터럴 '#' 문자 (포장 뒤에 붙음)
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # ([\d-]*)?               [Group 2] (옵션) 일련번호: 숫자 또는 하이픈이 올 수 있음 (예: 1, 12-1). 없으면 빈값.
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # ([가-힣a-zA-Z\s\/]+?)    [Group 3] 국가: 한글, 영문, 공백, 슬래시. '?'(Non-greedy)를 써서 뒤에 오는 수량/설명 전까지만 잡음.
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # (\d+)U?                  [Group 4] (옵션) 수량: 숫자+U 형태. 수량이 생략될 수 있어 옵션 처리.
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # (\(.*?\))?              [Group 5] (옵션) 설명: 괄호 (...) 안의 내용 (예: (수출용)). 없어도 됨.
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # ([X\d,\s,a-fA-F]+)      [[Group 6] 배치번호: 'X', 숫자, 쉼표, 공백 등이 포함된 문자열 (예: X12345, X12346)

                    pattern = r"(포장)#\s*([\d-]*)?\s*([가-힣a-zA-Z\s\/]+?)\s*(\d+U)?\s*(\(.*?\))?\s*([X\d,\s,a-fA-F]+)"
                    p_match = re.search(pattern, text)

                    if p_match:
                        all_plans.append({
                            "계획연도": year,
                            "일련번호": p_match.group(2) if p_match.group(2) else "",
                            "유형": "포장",
                            "국가": p_match.group(3),  # 미국
                            "단위(U)": p_match.group(4),  # 50, 100
                            "설명": p_match.group(5),  # 미국 수출용 포장
                            "시작일": start_date,
                            "종료일": end_date,
                            "배치번호": p_match.group(6)
                        })

    # CSV 저장
    if all_plans:
        df = pd.DataFrame(all_plans)
        # 컬럼 순서 정렬
        columns = ["계획연도", "일련번호", "유형", "국가", "단위(U)", "설명", "시작일", "종료일", "배치번호"]
        df = df[columns]

        output_path = "parsed_production_plan.csv"
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        return df
    else:
        return pd.DataFrame()

# 날짜 계산 로직 분리
def get_merged_date(ws, cell, date_row, current_col, year, month):
    start_col = current_col
    end_col = current_col

    # 병합 여부 확인
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            start_col = merged.min_col
            end_col = merged.max_col
            break

    try:
        # 날짜 셀 값 가져오기
        start_val = ws.cell(row=date_row, column=start_col).value
        end_val = ws.cell(row=date_row, column=end_col).value

        if not start_val or not end_val: return None, None

        start_date = start_val.replace(year=year, month=month)
        end_date = end_val.replace(year=year, month=month)
        return start_date, end_date
    except Exception as e:
        return None, None

# 실행 테스트
df_result = parse_production_schedule("product_plan.xlsx")
print(df_result)