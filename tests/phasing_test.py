import re
import pytest
try:
    import openpyxl
except Exception:
    openpyxl = None  # type: ignore
    pytest.skip("openpyxl is not installed; skipping phasing_test", allow_module_level=True)


def find_unmatched_patterns(file_path):
    """
    엑셀 파일을 순회하며 '생산' 또는 '포장#' 키워드는 있지만,
    정해진 정규 표현식에 매칭되지 않는(실패한) 셀 값을 찾아 반환합니다.
    """
    semi_records, pack_records = 0, 0
    wb = openpyxl.load_workbook(file_path, data_only=True)
    unmatched_records = []

    # [검증 대상 정규 표현식]
    # 1. 반제품 패턴 (생산)
    # 기존: 생산 / #1-1 / 국가 / 50U / X12345
    regex_semi = r"생산\s+(#[\d-]+)\s*([가-힣a-zA-Z\s\/]+)\s*(\d+U)\s*(\(.*?\))?\s*(X[\d\s,a-fA-F]+)?"


    # 2. 포장 패턴 (포장)
    # 기존: 포장#1 / 국가 / 50U / (설명) / X12345, X12346
    regex_pack = r"(포장)#\s*([\d-]*)?\s*([가-힣a-zA-Z\s\/]+?)\s*(\d+U)?\s*(\(.*\))?\s*([XZ\d,\s,a-fA-F]+)"

    print(f"[*] 데이터 검증을 시작합니다...")

    for sheet_name in wb.sheetnames:
        # 시트 필터링
        if not re.match(r'\d{6}', sheet_name): continue
        if not re.match(r'(2025\d+)', sheet_name): continue

        ws = wb[sheet_name]

        # 그리드 순회 (5주 ~ 6주 가정)
        for week_idx in range(6):
            date_row = 5 + (week_idx * 6)

            # 행 정의
            semi_rows = [date_row + 1, date_row + 2]  # 반제품
            pack_rows = [date_row + 3, date_row + 4]  # 포장

            # 빈 주차 확인
            if not ws.cell(row=date_row, column=5).value and not ws.cell(row=date_row, column=11).value:
                break

            # 1. 반제품 미매칭 검사
            for r in semi_rows:
                for c in range(5, 12):
                    cell = ws.cell(row=r, column=c)
                    if not cell.value: continue

                    text = str(cell.value).strip()

                    # [수정] 키워드가 있는 경우에만 '반제품 계획'으로 간주하고 카운트
                    if "생산" in text:
                        semi_records += 1  # <--- 여기로 이동
                        if not re.search(regex_semi, text):
                            unmatched_records.append({...})

            # 2. 포장 미매칭 검사
            for r in pack_rows:
                for c in range(5, 12):
                    cell = ws.cell(row=r, column=c)
                    if not cell.value: continue

                    text = str(cell.value).strip()

                    # [수정] 키워드가 있는 경우에만 '포장 계획'으로 간주하고 카운트
                    if "포장#" in text:
                        pack_records += 1  # <--- 여기로 이동
                        if not re.search(regex_pack, text):
                            unmatched_records.append({...})


    return semi_records, pack_records, unmatched_records


# --- 실행 및 결과 출력 ---
if __name__ == "__main__":
    # 파일 경로를 입력하세요
    target_file = "product_plan.xlsx"

    try:
        results = find_unmatched_patterns(target_file)

        if results:
            print(f"\n[!] 총 {results[0]}+{results[1]} 중 {len(results[2])}건의 미매칭 패턴을 발견했습니다.\n")
            print("-" * 80)
            print(f"{'Type':<15} | {'Sheet':<10} | {'Cell':<6} | {'Raw Text'}")
            print("-" * 80)

            for item in results[2]:
                print(f"{item['Type']:<15} | {item['Sheet']:<10} | {item['Cell']:<6} | {item['Raw_Text']}")

            print("-" * 80)
            print("위 텍스트들을 복사하여 정규 표현식 개선에 활용하세요.")
        else:
            print("\n[OK] 모든 '생산' 및 '포장#' 데이터가 정상적으로 파싱되었습니다.")

    except FileNotFoundError:
        print("파일을 찾을 수 없습니다. 경로를 확인해주세요.")
    except Exception as e:
        print(f"오류 발생: {e}")