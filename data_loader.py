import pandas as pd
import openpyxl
import os
import re

# 데이터 파일 경로 설정
# 추후 DB 연결 시 삭제
DATA_DIR = "data"

# 파일 종류별 전처리 로직
def _process_product_info(df):
    """
    자재 정보의 뼈대를 잡는 파일을 처리
    """
    # 컬럼 매핑
    mapping = {
        "자재 유형": "product_type", "플랜트": "plant_code", "자재": "product_id", "자재 내역": "description",
        "기본 단위": "base_unit", "플랜트별 자재 상태": "plant_status", "생산 저장 위치": "prod_storage_loc",
        "EP 저장 위치": "ep_storage_loc", "잔여 유효 기간": "remaining_shelf_life_days",
        "총 셀프 라이프": "total_shelf_life_days", "검사설정": "inspection_setting",
    }
    df.rename(columns=mapping, inplace=True)

    # 필수 컬럼 확인(필요 시 추가)
    # if "product_id" not in df.columns or "delivering_plant" not in df.columns:
    #     raise ValueError("필수 컬럼(자재, 플랜트)이 누락되었습니다.")

    # 필수 PK가 없으면 드랍
    df.dropna(subset=["product_id"], inplace=True)

    # 중복 제거 (Product ID 기준)
    df.drop_duplicates(subset=["product_id"], keep="last", inplace=True)

    # 필수 컬럼만 남기기 (데이터 제거)
    valid_cols = [c for c in mapping.values() if c in df.columns]
    return df[valid_cols]


def _process_attachment_info(df):
    """
    기존 product 테이블에 외부 착인 정보를 추가함
    """
    mapping = {
        "자재": "product_id",
        "외부착인": "is_attachment",
    }
    df.rename(columns=mapping, inplace=True)

    # 로직 적용(필요 시 추가)
    # df['is_attachment'] = df['is_attachment'].map({'X': "No", None: "Yes"})

    # 필수 컬럼만 남기기
    return df[mapping.values()]  # 필요한 컬럼만 리턴


def _process_edition_info(df):
    """
    기존 product 테이블에 에디션 정보를 추가함
    """
    mapping = {
        "자재": "product_id",
        "자재그룹": "product_group",
        "자재그룹 내역": "product_group_description",
        "Edition No.": "edition_no",
    }
    df.rename(columns=mapping, inplace=True)
    return df[mapping.values()]  # 필요한 컬럼만 리턴

def _process_vendor_info(df):
    """
    공급업체 정보 추가
    """
    mapping = {
        "공급업체": "vendor_id",
        "공급업체 이름": "vendor_name",
        "구매 조직": "purchase_org",
        "오더 통화": "order_currency"
    }
    df.rename(columns=mapping, inplace=True)
    return df[mapping.values()]  # 필요한 컬럼만 리턴


def _process_bom_info(df):
    """[BOM] 자재 명세서 (PK 없음 / 단순 리스트)"""
    mapping = {
        "자재번호(Root)": "product_id",
        "기준 수량": "standard_qty",
        "레벨": "level",
        "상위자재": "parent_product_id",
        "구성요소": "component_product_id",
        "구성부품수량": "component_qty",
    }
    df.rename(columns=mapping, inplace=True)
    # 데이터 타입 보정 (Agent가 Join할 때 중요)
    # 외래키(FK) 역할을 하는 컬럼들은 문자열로 통일해줘야 나중에 Join이 잘 됨
    for col in ["product_id", "parent_product_id", "component_product_id"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    return df[mapping.values()]


def _process_material_ledger_info(df):
    """[material_ledger] 자재수불부(기간별 흐름)"""
    mapping = {
        "자재": "product_id",
        "Cncy": "currency",

        # material_ledger
        "표준원가": "standard_price",
        "기초(수량)": "opening_qty",
        "기초(금액)합계": "opening_amount",

        "구매입고(수량)": "receipt_purchase_qty",
        "구매입고(금액)": "receipt_purchase_price",
        "구매입고(가격차이)": "receipt_purchase_price_diff",

        "생산입고(수량)": "production_receipt_qty",
        "생산입고(금액)": "production_receipt_price",
        "생산입고(가격차이)": "production_receipt_price_diff",

        "기타입고(수량)": "other_receipt_qty",
        "기타입고(금액)": "other_receipt_price",
        "기타입고(가격차이)": "other_receipt_price_diff",

        "가격 차이": "price_diff",

        "입고(수량)합계": "total_receipt_qty",
        "입고(금액)합계": "total_receipt_amount",
        "입고(가격차이)합계": "total_receipt_price_diff",

        "생산출고(수량)": "production_issue_qty",
        "생산출고(금액)": "production_issue_price",
        "생산출고(가격차이)": "production_issue_price_diff",

        "코스트센터출고(수량)": "cost_center_issue_qty",
        "코스트센터출고(금액)": "cost_center_issue_price",
        "코스트센터출고(가격차이)": "cost_center_issue_price_diff",

        "기타출고(수량)": "other_issue_qty",
        "기타출고(금액)": "other_issue_price",
        "기타출고(가격차이)": "other_issue_price_diff",

        "소비(수량)합계": "total_issue_qty",
        "소비(금액)합계": "total_issue_amount",
        "소비(가격차이)합계": "total_issue_price_diff",

        # "총 가격 차이": "total_price_diff",

        "기말(수량)": "closing_qty",
        "기말(금액)합계": "closing_amount",
    }

    df.rename(columns=mapping, inplace=True)
    # 데이터 타입 보정 (Agent가 Join할 때 중요)
    # 외래키(FK) 역할을 하는 컬럼들은 문자열로 통일해줘야 나중에 Join이 잘 됨
    if "product_id" in df.columns:
        df["product_id"] = df["product_id"].astype(str).str.strip()

    return df[mapping.values()]


def _get_merged_date(ws, cell, date_row, current_col, year, month):
    """[Helper] 병합된 셀의 범위를 인식하여 날짜 계산"""
    start_col = current_col
    end_col = current_col

    # 병합 셀 확인
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            start_col = merged.min_col
            end_col = merged.max_col
            break
    try:
        start_val = ws.cell(row=date_row, column=start_col).value
        end_val = ws.cell(row=date_row, column=end_col).value

        if not start_val or not end_val: return None, None

        # 엑셀의 날짜 데이터가 datetime 객체인 경우 replace 사용
        # 정수형 날짜라면 datetime으로 변환하는 로직 추가
        start_date = start_val.replace(year=year, month=month)
        end_date = end_val.replace(year=year, month=month)
        return start_date, end_date
    except Exception:
        return None, None

def _process_production_plan(uploaded_file):
    """
    [Plan] 생산 계획 파싱 (OpenPyXL 사용)
    주의: 이 함수는 DataFrame이 아닌 '파일 객체'를 입력받습니다.
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    all_plans = []

    for sheet_name in wb.sheetnames:
        # 시트 필터링 (YYYYMM 형식 & 2025년 이후)
        if not re.match(r'\d{6}', sheet_name): continue
        if not re.match(r'(2025\d+)', sheet_name): continue

        ws = wb[sheet_name]
        header_val = ws['E3'].value
        if not header_val: continue

        # E3 셀이 날짜 형식이면 month/year 속성 사용
        try:
            month, year = header_val.month, header_val.year
        except AttributeError:
            # 날짜 형식이 아닐 경우 예외 처리 추가 바람
            continue

        cols_range = range(5, 12)  # E~K (일~토)

        for week_idx in range(6):
            date_row = 5 + (week_idx * 6)
            semi_rows = [date_row + 1, date_row + 2]
            pack_rows = [date_row + 3, date_row + 4]

            # 빈 주차 확인
            if not ws.cell(row=date_row, column=5).value and not ws.cell(row=date_row, column=11).value:
                break

            # 반제품 생산 계획
            for r in semi_rows:
                for c in cols_range:
                    cell = ws.cell(row=r, column=c)
                    if not cell.value: continue
                    text = str(cell.value).strip()

                    if "생산" not in text: continue

                    start_date, end_date = _get_merged_date(ws, cell, date_row, c, year, month)
                    if not start_date: continue

                    # 생산                      리터럴 '생산' 문자열과 매칭
                    # \s+                      공백 문자(스페이스, 탭 등) 1개 이상 (필수 구분자)
                    # (#[\d-]+)                [Group 1] 일련번호: '#'으로 시작하고 숫자나 하이픈(-)이 옴
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # ([가-힣a-zA-Z\s\/]+?)     [Group 2] 국가/유형: 한글, 영문, 공백, 슬래시(/) 포함
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # (\d+)U?                  [Group 3] (옵션)수량: 숫자+U 형태
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # (\(.* ?\))?              [Group 4] (옵션) 설명: 괄호 (...)로 묶인 추가 정보
                    # (X[\d\s,a-fA-F]+)?       [Group 5] (옵션) 배치번호: 'X'로 시작하며 숫자, 문자 포함

                    # Regex: 생산 #1-1 유럽 50U (설명) X12345
                    pattern = r"생산\s+(#[\d-]+)\s*([가-힣a-zA-Z\s\/]+?)\s*(\d+)U?\s*(\(.*?\))?\s*(X[\d\s,a-fA-F]+)?"
                    p_match = re.search(pattern, text)

                    if p_match:
                        all_plans.append({
                            "serial_no": p_match.group(1),
                            "material_type": "반제품",
                            "country": p_match.group(2),
                            "quantity": p_match.group(3),
                            "remark": p_match.group(4) if p_match.group(4) else "",
                            "start_date": start_date,
                            "end_date": end_date,
                            "batch_no": p_match.group(5)
                        })

            # 포장 계획
            for r in pack_rows:
                for c in cols_range:
                    cell = ws.cell(row=r, column=c)
                    if not cell.value: continue
                    text = str(cell.value).strip()

                    if "포장#" not in text: continue

                    start_date, end_date = _get_merged_date(ws, cell, date_row, c, year, month)
                    if not start_date: continue

                    # 포장                      [Group 1] 리터럴 '포장'
                    # #                        리터럴 '#' 문자 (포장 뒤에 붙음)
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # ([\d-]*)?                [Group 2] (옵션) 일련번호: 숫자 또는 하이픈이 올 수 있음
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # ([가-힣a-zA-Z\s\/]+?)     [Group 3] 국가: 한글, 영문, 공백, 슬래시. '?'(Non-greedy)를 써서 뒤에 오는 수량/설명 전까지만 잡음.
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # (\d+)U?                  [Group 4] (옵션) 수량: 숫자+U 형태.
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # (\(.*?\))?               [Group 5] (옵션) 설명: 괄호 (...) 안의 내용
                    # \s*                      공백 문자 0개 이상 (유연하게 처리)
                    # ([X\d,\s,a-fA-F]+)       [Group 6] 배치번호: 'X', 숫자, 쉼표, 공백 등이 포함된 문자열

                    # Regex: 포장#1 미국 50U (설명) X12345
                    # pattern = r"(포장)#\s*([\d-]*)?\s*([가-힣a-zA-Z\s\/]+?)\s*(\d+U)?\s*(\(.*?\))?\s*([X\d,\s,a-fA-F]+)"
                    pattern = r"(포장)#\s*([\d-]*)?\s*([가-힣a-zA-Z\s\/]+?)\s*(\d+)U?\s*(\(.*\))?\s*([XZ\d,\s,a-fA-F]+)"
                    p_match = re.search(pattern, text)

                    if p_match:
                        qty_val = p_match.group(4).replace("U", "") if p_match.group(4) else 0
                        all_plans.append({
                            "serial_no": p_match.group(2) if p_match.group(2) else "",
                            "material_type": "포장",
                            "country": p_match.group(3),
                            "quantity": qty_val,
                            "remark": p_match.group(5) if p_match.group(5) else "",
                            "start_date": start_date,
                            "end_date": end_date,
                            "batch_no": p_match.group(6)
                        })

    if all_plans:
        return pd.DataFrame(all_plans)
    else:
        return pd.DataFrame(columns=TABLE_SCHEMA["manufacture_plan"])



# 프로세서 등록 (Registry)
FILE_PROCESSORS = {
    # (UI표시 이름) : (타겟 테이블, 처리 함수, 병합 전략)
    "자재 정보": ("product", _process_product_info, "UPSERT_ROWS", "product_id", ),
    "자재 외부 착인 여부": ("product", _process_attachment_info, "EXTEND_COLUMNS", "product_id", ),
    "자재 에디션 숫자": ("product", _process_edition_info, "EXTEND_COLUMNS", "product_id", ),
    "공급업체 목록": ("vendor",_process_vendor_info,"UPSERT_ROWS", "vendor_id", ),
    "BOM": ("bom", _process_bom_info, "REPLACE_ALL", None),
    "생산 계획": ("manufacture_plan", _process_production_plan, "REPLACE_ALL", None),
    "자재수불부": ("material_ledger", _process_material_ledger_info, "REPLACE_ALL", None),
}


def save_uploaded_file_by_type(uploaded_file, source_type):
    """
    # xlsx 데이터 통합 저장 로직 (Merge & Load)
    """
    if source_type not in FILE_PROCESSORS:
        return False, f"지원하지 않는 파일 형식입니다."

    target_table, processor_func, strategy, pk_col = FILE_PROCESSORS[source_type]
    # 기존 데이터를 불러오기 위한 경로 설정
    file_path = os.path.join(DATA_DIR, f"{target_table}.csv")

    try:
        # 엑셀 파일 로드 및 전처리
        if source_type == "생산 계획":
            new_df = processor_func(uploaded_file)  # openpyxl은 파일 객체를 직접 필요로 함
        else:
            raw_df = pd.read_excel(uploaded_file)
            new_df = processor_func(raw_df)

        # 단순 교체 전략 (BOM 등 PK가 없는 경우)
        if strategy == "REPLACE_ALL":
            # 복잡한 병합 로직 없이 바로 저장 (덮어쓰기)
            if not os.path.exists(DATA_DIR):
                os.makedirs(DATA_DIR)

            new_df.to_csv(file_path, index=False)
            return True, f"[{source_type}] 저장 완료. {len(new_df)}건)"

        # PK가 필요한 전략인데 PK가 None이면 에러 처리
        if strategy in ["UPSERT_ROWS", "EXTEND_COLUMNS"] and not pk_col:
            return False, "설정 오류: 해당 전략은 식별자(PK)가 필요합니다."

        # 새로운 데이터에 대해 숫자로 들어오든 문자로 들어오든 무조건 str로 맞추고 공백을 날립니다.
        new_df[pk_col] = new_df[pk_col].astype(str).str.strip()

        # 기존 데이터가 존재하면, pk_col을 전처리
        if os.path.exists(file_path):
            current_df = pd.read_csv(file_path)
            current_df[pk_col] = current_df[pk_col].astype(str).str.strip()
        else:
            current_df = pd.DataFrame()

        if current_df.empty:
            final_df = new_df
        else:
            # 병합을 위해 pk_col을 인덱스로 설정
            current_df.set_index(pk_col, inplace=True)
            new_df.set_index(pk_col, inplace=True)

            # 병합 전략 실행
            if strategy == "UPSERT_ROWS":
                # 행 중심 병합: 새로운 컬럼이 있으면 추가하되, 주로 행(Row) 데이터를 최신화함

                # 컬럼 동기화 (새 파일에 없는 컬럼은 유지)
                final_df = current_df.combine_first(new_df)
                # update로 덮어쓰기 (new_df의 값이 우선)
                final_df.update(new_df)

            elif strategy == "EXTEND_COLUMNS":
                # 열 중심 병합: 기존 행은 건드리지 않고, 매칭되는 ID에 대해서만 값을 갱신/추가
                # 스키마 확장: 새 파일에 있는 컬럼이 기존에 없으면 빈 컬럼 추가
                for col in new_df.columns:
                    if col not in current_df.columns:
                        current_df[col] = pd.NA

                # 값 업데이트: 인덱스(ID)가 일치하는 곳에 값 덮어쓰기
                # update는 교집합(index가 양쪽에 모두 존재)에 대해서만 작동함
                current_df.update(new_df)
                final_df = current_df
            else:
                final_df = pd.DataFrame()
            # 병합 종료 후 인덱스를 다시 컬럼으로 변환
            final_df.reset_index(inplace=True)

        # 저장
        if not os.path.exists(DATA_DIR):
            os.makedirs(DATA_DIR)
        final_df.to_csv(file_path, index=False)

        return True, f"[{source_type}] 처리 완료. (전략: {strategy}, 총 {len(final_df)}건)"

    except Exception as e:
        return False, f"오류 발생: {str(e)}"


def load_master_data():
    """
    실제 CSV 파일을 로드하여 데이터프레임 딕셔너리로 반환
    CSV를 로드하되, LLM용(Secure) 데이터와 UI용(Mapping) 데이터를 분리하여 반환
    """
    db = {}

    try:
        # TABLE_SCHEMA에 정의된 테이블들을 순회하며 로드
        for table_name in TABLE_SCHEMA.keys():
            file_path = os.path.join(DATA_DIR, f"{table_name}.csv")

            if os.path.exists(file_path):
                # 원본 데이터 그대로 로드 (Description 포함)
                df = pd.read_csv(file_path)

                # 데이터 타입 보정 (예: 날짜, 숫자 등)이 필요하면 여기서 수행
                db[table_name] = df
            else:
                # 파일이 없으면 빈 DataFrame 생성
                db[table_name] = pd.DataFrame(columns=TABLE_SCHEMA.get(table_name, []))

        return db

    except Exception as e:
        print(f"데이터 로드 중 오류: {e}")
        return {}