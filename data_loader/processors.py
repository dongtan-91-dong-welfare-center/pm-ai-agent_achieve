import pandas as pd

try:
    import openpyxl

    _HAS_OPENPYXL = True
except Exception:
    # openpyxl may not be available in minimal test environments
    openpyxl = None  # type: ignore
    _HAS_OPENPYXL = False
import re
from typing import List, Union


# =============================================================================
# 1. Helper Functions (Data Normalization)
# =============================================================================

def get_merged_date(ws, cell, date_row, current_col, year, month):
    """병합된 셀의 범위를 인식하여 날짜 계산"""
    start_col = current_col
    end_col = current_col

    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            start_col = merged.min_col
            end_col = merged.max_col
            break
    try:
        start_val = ws.cell(row=date_row, column=start_col).value
        end_val = ws.cell(row=date_row, column=end_col).value

        if not start_val or not end_val: return None, None

        start_date = start_val.replace(year=year, month=month)
        end_date = end_val.replace(year=year, month=month)
        return start_date, end_date
    except Exception:
        return None, None


def _clean_str_col(series: pd.Series) -> pd.Series:
    """[공통] 문자열 기본 정제: 변환 -> 공백 제거 -> 소수점(.0) 제거"""
    return series.astype(str).str.strip().str.replace(r'\.0$', '', regex=True)


def _normalize_product_id(df: pd.DataFrame, cols: List[str] = ['product_id']) -> pd.DataFrame:
    """[규칙 적용] 품목코드: 7자리 숫자 (Zero Padding)"""
    for col in cols:
        if col in df.columns:
            # 예: '123' -> '0000123'
            df[col] = _clean_str_col(df[col]).str.zfill(7)
    return df


def _normalize_batch_no(df: pd.DataFrame, cols: List[str] = ['batch_no']) -> pd.DataFrame:
    """[규칙 적용] 배치번호: 7자리 숫자 + 1자리 알파벳 = 총 8자리 (Zero Padding)"""
    for col in cols:
        if col in df.columns:
            # 예: '2500N' -> '0002500N'
            df[col] = _clean_str_col(df[col]).str.zfill(8)
    return df


def _normalize_generic_id(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    """[일반] 기타 ID(PO, 업체코드 등): 문자열 변환 및 소수점 제거만 수행"""
    for col in cols:
        if col in df.columns:
            df[col] = _clean_str_col(df[col])
    return df


# =============================================================================
# 2. File Processors (Data Loading Logic)
# =============================================================================

def process_product_info(df):
    """[Product] 자재 정보"""
    mapping = {
        "자재 유형": "product_type", "플랜트": "plant_code", "자재": "product_id", "자재 내역": "description",
        "기본 단위": "base_unit", "플랜트별 자재 상태": "plant_status", "생산 저장 위치": "prod_storage_loc",
        "EP 저장 위치": "ep_storage_loc", "잔여 유효 기간": "remaining_shelf_life", "총 셀프 라이프": "total_shelf_life",
        "검사설정": "inspection_setting",
    }
    df.rename(columns=mapping, inplace=True)
    df.dropna(subset=["product_id"], inplace=True)
    df.drop_duplicates(subset=["product_id"], keep="last", inplace=True)

    # 정규화: 품목코드
    df = _normalize_product_id(df, ["product_id"])

    valid_cols = [c for c in mapping.values() if c in df.columns]
    return df[valid_cols]


def process_edition_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Product] 에디션 정보"""
    mapping = {
        "자재": "product_id",
        "자재그룹": "product_group",
        "자재그룹 내역": "product_group_name",
        "Edition No.": "edition_no",
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    return df[list(mapping.values())]


def process_attachment_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Product] 외부 착인 정보"""
    mapping = {
        "자재": "product_id",
        "외부착인": "is_attachment",
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_bom_info(df: pd.DataFrame) -> pd.DataFrame:
    """[BOM] 자재 명세서"""
    mapping = {
        "자재번호(Root)": "root_product_id",
        "기준 수량": "standard_qty",
        "레벨": "level",
        "상위자재": "parent_product_id",
        "구성요소": "component_product_id",
        "구성부품수량": "component_qty",
    }
    df.rename(columns=mapping, inplace=True)
    # 정규화: 관련된 모든 품목코드 컬럼
    df = _normalize_product_id(df, ["root_product_id", "parent_product_id", "component_product_id"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_vendor_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Vendor] 공급업체 정보"""
    mapping = {
        "Fix": "is_fixed_vendor",
        "플랜트": "plant_code",
        "공급업체": "vendor_id",
        "업체명": "vendor_name",
        "자재번호": "product_id",
        "유효시작일": "valid_from",
        "유효만료일": "valid_to",
        "제조자": "manufacturer_id",
        "제조자명": "manufacturer_name",
        "구매그룹": "purchasing_group",
        "구매그룹명": "purchasing_group_name",
        "발주단위": "order_unit",
        "기본단위": "base_unit",
        "단가": "unit_price",
        "통화": "currency",
        "가격단위": "price_unit"
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    df = _normalize_generic_id(df, ["vendor_id", "manufacturer_id"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_overage_rule_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Overage] 오버리지 기준"""
    mapping = {
        "자재": "product_id",
        "포장재코드": "packing_code",
        "투입량 범위(FROM)": "range_from",
        "투입량 범위(TO)": "range_to",
        "오버리지 수량(절대값)": "overage_abs_qty",
        "오버리지 비율(%)": "overage_rate",
        "올림 소수점 자릿수": "rounding_decimal",
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_material_ledger_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Material_Ledger] 자재수불부"""
    mapping = {
        "자재": "product_id",
        "Cncy": "currency",
        "표준원가": "standard_price",
        "기초(수량)": "opening_qty",
        "기초(금액)합계": "opening_amount",
        "구매입고(수량)": "receipt_purchase_qty",
        "구매입고(금액)": "receipt_purchase_price",
        "구매입고(가격차이)": "receipt_purchase_price_diff",
        "생산입고(수량)": "production_receipt_qty",
        "생산입고(금액)": "production_receipt_price",
        "생산입고(가격차이)": "production_receipt_price_diff",
        "기타입고(수량)": "other_receipt_qty",
        "기타입고(금액)": "other_receipt_price",
        "기타입고(가격차이)": "other_receipt_price_diff",
        "가격 차이": "price_diff",
        "입고(수량)합계": "total_receipt_qty",
        "입고(금액)합계": "total_receipt_amount",
        "입고(가격차이)합계": "total_receipt_price_diff",
        "생산출고(수량)": "production_issue_qty",
        "생산출고(금액)": "production_issue_price",
        "생산출고(가격차이)": "production_issue_price_diff",
        "코스트센터출고(수량)": "cost_center_issue_qty",
        "코스트센터출고(금액)": "cost_center_issue_price",
        "코스트센터출고(가격차이)": "cost_center_issue_price_diff",
        "코스트센터출고(조정)": "cost_center_issue_adjustment",
        "기타출고(수량)": "other_issue_qty",
        "기타출고(금액)": "other_issue_price",
        "기타출고(가격차이)": "other_issue_price_diff",
        "소비(수량)합계": "total_issue_qty",
        "소비(금액)합계": "total_issue_amount",
        "소비(가격차이)합계": "total_issue_price_diff",
        "가격 차이": "total_price_diff",
        "기말(수량)": "closing_qty",
        "기말(금액)합계": "closing_amount",
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_purchase_transaction_history_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Purchase_Transaction_History] 구매 상세 내역"""
    mapping = {
        "구매문서번호": "po_id",
        "구매품목": "po_item_no",
        "입고일자": "receipt_date",
        "이동유형": "movement_type",
        "자재코드": "product_id",
        "오더수량": "order_qty",
        "Info.Rec.수정일": "info_rec_date",
        "OLD값": "old_price",
        "NEW값": "new_price",
        "마스터단가": "master_price",
        "마스터단가통화": "master_price_currency",
        "오더단가": "order_price",
        "통화": "order_currency",
        "단가단위수량": "price_unit",
        "입고수량": "received_quantity",
        "입고금액(발주통화)": "received_value_local_currency",
        "제판비": "printing_plate_cost",
        "동판비": "copper_plate_cost",
        "입고총금액(발주통화)": "total_received_value_local_currency",
        "입고총금액(원화)": "total_received_value_krw",
        "구매업체": "vendor_id"
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    df = _normalize_generic_id(df, ["po_id", "vendor_id"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_good_receipt_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Good_Receipt] 입고 이력"""
    mapping = {
        "작업일시": "work_datetime",
        "구매 문서 번호": "po_id",
        "구매 문서 품목 번호": "item_no",
        "자재 번호": "product_id",
        "배치번호": "batch_no",
        "제조번호": "manufacturing_no",
        "제조일자": "manufacturing_date",
        "유효일자": "expiration_date",
        "종입고수량": "receipt_qty",
        "검사로트번호": "inspection_lot_no",
    }
    df.rename(columns=mapping, inplace=True)

    # 정규화: Product(7자리), Batch(8자리), PO(일반)
    df = _normalize_product_id(df, ["product_id"])
    df = _normalize_batch_no(df, ["batch_no"])
    df = _normalize_generic_id(df, ["po_id"])

    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_purchase_order_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Purchase_Order] 구매 오더"""
    mapping = {
        "구매 문서": "po_id",
        "품목": "item_no",
        "공급업체": "vendor_id",
        "자재": "product_id",
        "구매 오더일": "po_date",
        "예정 수량": "schedule_qty",
        "GR 수량": "received_qty",
        "납품일": "delivery_date"
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    df = _normalize_generic_id(df, ["po_id", "vendor_id"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_batch_stock_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Batch_Stock] 유효 기한 정보"""
    mapping = {
        "자재": "product_id",
        "제조일": "manufacture_date",
        "유효 기한": "expiration_date",
        "배치": "batch_no",
        "자재 그룹": "material_group",
        "가용": "available_qty",
        "품질 검사": "quality_inspection_qty",
        "보류재고": "blocked_stock",
        "가용재고 값": "available_stock_value",
        "품질검사재고 값": "quality_inspection_stock_value",
        "보류재고 값": "blocked_stock_value",
        "입고일": "receipt_date",
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    df = _normalize_batch_no(df, ["batch_no"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_warehouse_stock_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Warehouse_Stock] 창고 재고"""
    mapping = {
        "자재": "product_id",
        "배치": "batch_no",
        "가용": "unrestricted_qty",
        "품질 검사": "inspection_qty",
        "사용 제한 재고": "restricted_stock",
        "보류재고": "blocked_qty",
        "반품": "return_qty",
        "평가 입고 보류 재고": "valuated_blocked_stock",
        "사용 용기": "issued_container",
        "운송 중 재고": "in_transit_stock",
        "운송 및 전송": "shipment_and_transfer",
        "재고 세그먼트": "stock_segment",
        "이전중 (플랜트)": "plant_transfer_in_progress",
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    df = _normalize_batch_no(df, ["batch_no"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_prod_plan_code_map_info(df: pd.DataFrame) -> pd.DataFrame:
    """[prod_plan_code_map] 생산품목코드 매핑"""
    mapping = {
        "자재": "product_id",
        "자재 내역": "description",
        "국가": "country",
        "유닛": "packing_unit",
    }
    df.rename(columns=mapping, inplace=True)
    df = _normalize_product_id(df, ["product_id"])
    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_non_conformance_info(df: pd.DataFrame) -> pd.DataFrame:
    """[Non_Conformance] 부적합 정보"""
    df.columns = df.columns.str.strip()

    mapping = {
        "자재": "product_id",
        "자재 내역": "description",
        "Plnt": "plant_code",
        "저장 위치": "storage_location",
        "이동 유형 텍스트": "movement_type_text",
        "이동 유형": "movement_type",
        "특별 재고": "special_stock",
        "자재 문서": "material_document",
        "자재 문서 항목": "material_document_item",
        "전기일": "posting_date",
        "증빙일": "document_date",
        "배치": "batch_no",
        "수량(입력 단위)": "entry_quantity",
        "입력단위": "entry_unit",
        "입력일": "entry_date",
        "입력 시간": "entry_time",
        "사용자 이름": "user_name",
        "오더": "order_id",
        "구매 오더": "purchase_order",
        "판매 오더": "sales_order",
        "판매 오더 품목": "sales_order_item",
        "문서 헤더 텍스트": "header_text",
        "이동지시자": "movement_indicator",
        "자재 수령인": "goods_recipient",
        "금액(현지 통화)": "amount_local_currency"
    }

    df.rename(columns=mapping, inplace=True)

    # 필터링
    if "header_text" in df.columns:
        df["header_text"] = df["header_text"].fillna("").astype(str).str.strip()
        df.loc[df["header_text"].str.lower() == "nan", "header_text"] = ""
        df = df[df["header_text"] == ""]

    # 정규화 적용
    df = _normalize_product_id(df, ["product_id"])
    df = _normalize_batch_no(df, ["batch_no"])
    df = _normalize_generic_id(df, ["purchase_order", "order_id", "sales_order"])

    cols = [c for c in list(mapping.values()) if c in df.columns]
    return df[cols]


def process_production_plan(uploaded_file) -> pd.DataFrame:
    """[Plan] 생산 계획 파싱"""
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl required.")

    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    all_plans = []

    for sheet_name in wb.sheetnames:
        if not re.match(r'\d{6}', sheet_name): continue
        if not re.match(r'(2025\d+)', sheet_name): continue

        ws = wb[sheet_name]
        header_val = ws['E3'].value
        if not header_val: continue

        try:
            month, year = header_val.month, header_val.year
        except AttributeError:
            continue

        cols_range = range(5, 12)

        for week_idx in range(6):
            date_row = 5 + (week_idx * 6)
            semi_rows = [date_row + 1, date_row + 2]
            pack_rows = [date_row + 3, date_row + 4]

            if not ws.cell(row=date_row, column=5).value and not ws.cell(row=date_row, column=11).value:
                break

            # 반제품
            for r in semi_rows:
                for c in cols_range:
                    cell = ws.cell(row=r, column=c)
                    if not cell.value: continue
                    text = str(cell.value).strip()
                    if "생산" not in text: continue

                    start_date, end_date = get_merged_date(ws, cell, date_row, c, year, month)
                    if not start_date: continue

                    pattern = r"생산\s+(#[\d-]+)\s*([가-힣a-zA-Z\s\/]+?)\s*(\d+)U?\s*(\(.*?\))?\s*(X[\d\s,a-fA-F]+)?"
                    p_match = re.search(pattern, text)
                    if p_match:
                        all_plans.append({
                            "serial_no": p_match.group(1),
                            "material_type": "반제품",
                            "country": p_match.group(2),
                            "packing_unit": p_match.group(3),
                            "remark": p_match.group(4) if p_match.group(4) else "",
                            "start_date": start_date,
                            "end_date": end_date,
                            "batch_no": p_match.group(5)
                        })

            # 포장
            for r in pack_rows:
                for c in cols_range:
                    cell = ws.cell(row=r, column=c)
                    if not cell.value: continue
                    text = str(cell.value).strip()
                    if "포장#" not in text: continue

                    start_date, end_date = get_merged_date(ws, cell, date_row, c, year, month)
                    if not start_date: continue

                    pattern = r"(포장)#\s*([\d-]*)?\s*([가-힣a-zA-Z\s\/]+?)\s*(\d+)U?\s*(\(.*\))?\s*([XZ\d,\s,a-fA-F]+)"
                    p_match = re.search(pattern, text)
                    if p_match:
                        qty_val = p_match.group(4).replace("U", "") if p_match.group(4) else 0
                        all_plans.append({
                            "serial_no": p_match.group(2) if p_match.group(2) else "",
                            "material_type": "포장",
                            "country": p_match.group(3),
                            "packing_unit": qty_val,
                            "remark": p_match.group(5) if p_match.group(5) else "",
                            "start_date": start_date,
                            "end_date": end_date,
                            "batch_no": p_match.group(6)
                        })

    if all_plans:
        return pd.DataFrame(all_plans)
    else:
        return pd.DataFrame()